#!/usr/bin/env python3
"""Erzeugt aus der ausgefüllten Erhebungsmappe den Konfigurationsentwurf.

    python3 tools/config_aus_erhebung.py \
        ../erhebung/SK-001_Erhebung_Schnittstelle_SAP.xlsx \
        -o config.yaml

Das Ergebnis ist ein Entwurf, keine fertige Konfiguration: Transformationen und
Quellpfade in epilot muss weiterhin ein Mensch setzen. Was in der Mappe fehlt,
wird als TODO-Kommentar übernommen statt geraten - eine geratene Angabe kostet
einen Fehlversuch in SAP.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

# Kodierungsangaben, wie sie in der Mappe stehen können -> Python-Name
KODIERUNGEN = {
    "utf-8": "utf-8", "utf8": "utf-8",
    "utf-8 mit bom": "utf-8", "utf8 mit bom": "utf-8",
    "windows-1252": "cp1252", "windows1252": "cp1252", "cp1252": "cp1252", "ansi": "cp1252",
    "iso-8859-1": "latin-1", "iso8859-1": "latin-1", "latin-1": "latin-1", "latin1": "latin-1",
}
TRENNZEICHEN = {"semikolon": ";", "komma": ",", "tabulator": "\t", "tab": "\t"}


def txt(zelle) -> str:
    return str(zelle.value).strip() if zelle.value is not None else ""


def lies_dateiformat(wb) -> tuple[dict, list[str]]:
    ws, fmt, offen = wb["1 Dateiformat"], {}, []
    antworten = {txt(ws.cell(r, 2)).lower(): txt(ws.cell(r, 3)) for r in range(6, 30)}

    roh = antworten.get("zeichenkodierung", "")
    if roh:
        schluessel = roh.lower().replace("_", "-")
        fmt["kodierung"] = KODIERUNGEN.get(schluessel, roh)
        fmt["bom"] = "bom" in schluessel
        if schluessel not in KODIERUNGEN:
            offen.append(f"Zeichenkodierung '{roh}' nicht erkannt - bitte prüfen")
    else:
        offen.append("Zeichenkodierung fehlt - DAS ist die Entscheidungszeile für die Werkzeugwahl")

    roh = antworten.get("trennzeichen", "")
    if roh:
        fmt["trennzeichen"] = TRENNZEICHEN.get(roh.lower(), roh[:1])
    else:
        offen.append("Trennzeichen fehlt")

    roh = antworten.get("zeilenende", "").upper()
    if "CRLF" in roh:
        fmt["zeilenende"] = "CRLF"
    elif "LF" in roh:
        fmt["zeilenende"] = "LF"
    else:
        offen.append("Zeilenende fehlt")

    roh = antworten.get("kopfzeile vorhanden", "").lower()
    if roh:
        fmt["kopfzeile"] = roh.startswith("j") or roh.startswith("y")

    roh = antworten.get("dezimaltrennzeichen", "").lower()
    if roh:
        fmt["dezimaltrennzeichen"] = "," if "komma" in roh else "."

    return fmt, offen


def lies_spalten(wb) -> tuple[list[dict], list[str]]:
    ws, spalten, offen = wb["2 Feldmapping"], [], []
    roh = []
    for r in range(9, ws.max_row + 1):
        name = txt(ws.cell(r, 5))          # Spaltenname in der CSV
        if not name:
            continue
        pos = txt(ws.cell(r, 4))
        roh.append({
            "position": int(pos) if pos.isdigit() else None,
            "name": name,
            "fachlich": txt(ws.cell(r, 2)),
            "laenge": txt(ws.cell(r, 9)),
            "pflicht": txt(ws.cell(r, 10)).upper().startswith("J"),
            "format": txt(ws.cell(r, 11)),
            "quelle": txt(ws.cell(r, 13)),
        })

    if not roh:
        return [], ["Blatt '2 Feldmapping' enthält keine Spaltennamen - Rücklauf fehlt noch"]

    ohne_pos = [z for z in roh if z["position"] is None]
    if ohne_pos:
        offen.append(f"{len(ohne_pos)} Spalten ohne Positionsangabe - Reihenfolge aus der Mappe übernommen")
    roh.sort(key=lambda z: (z["position"] is None, z["position"] or 0))

    for z in roh:
        eintrag: dict = {"name": z["name"]}
        if z["quelle"]:
            eintrag["quelle"] = z["quelle"]
        else:
            eintrag["quelle"] = f"TODO_{z['name'].lower()}"
            offen.append(f"Spalte '{z['name']}' ({z['fachlich']}): epilot-Quelle fehlt")
        if z["pflicht"]:
            eintrag["pflicht"] = True
        if z["laenge"].isdigit():
            eintrag["max_laenge"] = int(z["laenge"])
        if z["format"]:
            eintrag["_format_laut_erhebung"] = z["format"]
            offen.append(f"Spalte '{z['name']}': Format '{z['format']}' - Transformation prüfen")
        spalten.append(eintrag)
    return spalten, offen


def lies_wertelisten(wb) -> tuple[dict, list[str]]:
    ws, listen, offen = wb["3 Wertelisten"], {}, []
    feld = None
    for r in range(6, ws.max_row + 1):
        if txt(ws.cell(r, 2)):
            feld = txt(ws.cell(r, 2)).lower().replace(" ", "_")
        klartext, sap, epi = txt(ws.cell(r, 3)), txt(ws.cell(r, 4)), txt(ws.cell(r, 5))
        if not feld or not klartext:
            continue
        if sap and epi:
            listen.setdefault(feld, {})[epi] = sap
        elif sap and not epi:
            offen.append(f"Werteliste '{feld}': SAP-Schlüssel '{sap}' für '{klartext}' hat keinen epilot-Wert")
    return listen, offen


def als_yaml(fmt, spalten, listen) -> str:
    import yaml
    entwurf = {
        "format": {**{"kodierung": "TODO", "trennzeichen": ";", "zeilenende": "CRLF",
                      "kopfzeile": True, "dezimaltrennzeichen": ","}, **fmt},
        "ablage": {"verzeichnis": "TODO", "dateiname": "TODO_%Y%m%d_%H%M%S.csv"},
        "epilot": {"schema": "netzanschluss_anfrage",
                   "query": "_schema:netzanschluss_anfrage AND uebertragungsstatus:bereit"},
        "wertelisten": listen,
        "pruefungen": [],
        "spalten": spalten,
    }
    return yaml.safe_dump(entwurf, allow_unicode=True, sort_keys=False, width=100)


def main(argv=None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("mappe", help="ausgefüllte Erhebungsmappe (.xlsx)")
    p.add_argument("-o", "--ausgabe", default="config.entwurf.yaml")
    a = p.parse_args(argv)

    wb = load_workbook(a.mappe, data_only=True)
    fmt, o1 = lies_dateiformat(wb)
    spalten, o2 = lies_spalten(wb)
    listen, o3 = lies_wertelisten(wb)
    offen = o1 + o2 + o3

    kopf = ["# Aus der Erhebungsmappe erzeugter Entwurf - vor dem Einsatz durchgehen.", "#"]
    if offen:
        kopf.append(f"# {len(offen)} offene Punkte:")
        kopf += [f"#   - {t}" for t in offen]
    else:
        kopf.append("# Keine offenen Punkte gefunden. Transformationen trotzdem prüfen.")
    kopf.append("")

    Path(a.ausgabe).write_text("\n".join(kopf) + als_yaml(fmt, spalten, listen), encoding="utf-8")

    print(f"Entwurf geschrieben: {a.ausgabe}")
    print(f"  Spalten: {len(spalten)}   Wertelisten: {len(listen)}   offene Punkte: {len(offen)}")
    for t in offen[:15]:
        print(f"  - {t}")
    if len(offen) > 15:
        print(f"  ... und {len(offen) - 15} weitere, siehe Kopf der Datei")
    return 1 if offen else 0


if __name__ == "__main__":
    sys.exit(main())
